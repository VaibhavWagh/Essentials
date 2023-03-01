#Create a list with username 

#Run the following command to get userwise usage information

#for i in `cat username`;do echo $i; find /home/SHARED/ANALYSIS/ -maxdepth 1 -user $i -type d ! -path . -exec du -chs --time {} + >> /home/SHARED/ANALYSIS/vaibhav/Usage/28_02_2023/$i".tsv" ;done;

#Keep all the above tsv files into a directory here "/home/vaibhav/Documents/Python/DU/Files/"
#Change the folder_path and run the script to get "Usage_information_userwise.xlsx"

import os
import openpyxl
import humanfriendly
from openpyxl.chart.label import DataLabelList


# Define the folder path containing the TSV files
folder_path = "/home/vaibhav/Documents/Python/DU/Files/"

# Get a list of all the TSV files in the folder
tsv_files = [f for f in os.listdir(folder_path) if f.endswith('.tsv')]

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Create a summary sheet
summary_sheet = workbook.active
summary_sheet.title = "Summary"
summary_sheet.cell(row=1, column=1).value = "User Name"
summary_sheet.cell(row=1, column=2).value = "Data Size (bytes)"
summary_sheet.cell(row=1, column=3).value = "Data & Time"
summary_sheet.cell(row=1, column=4).value = "Folder Path"
#summary_sheet.cell(row=1, column=5).value = "Percentage"

# Iterate over the TSV files and add them as worksheets to the workbook
for tsv_file in tsv_files:
    # Get the worksheet name from the filename
    sheet_name = os.path.splitext(tsv_file)[0]
    worksheet = workbook.create_sheet(title=sheet_name)
    with open(os.path.join(folder_path, tsv_file), 'r') as f:
        # Add column headers
        worksheet.cell(row=1, column=1).value = "Data Size"
        worksheet.cell(row=1, column=2).value = "Data & Time"
        worksheet.cell(row=1, column=3).value = "Folder Path"

        # Add data from the TSV file
        row_num = 2
        total_size = 0
        for line in f:
            data = line.strip().split('\t')
            # Convert data size to bytes using humanfriendly library
            data_size_bytes = humanfriendly.parse_size(data[0])
            total_size += data_size_bytes
            worksheet.cell(row=row_num, column=1).value = data[0]
            worksheet.cell(row=row_num, column=2).value = data[1]
            worksheet.cell(row=row_num, column=3).value = data[2]
            row_num += 1


        # Add total row to the summary sheet
        summary_sheet.cell(row=summary_sheet.max_row+1, column=1).value = sheet_name
        summary_sheet.cell(row=summary_sheet.max_row, column=2).value = total_size
        summary_sheet.cell(row=summary_sheet.max_row, column=3).value = data[1]
        summary_sheet.cell(row=summary_sheet.max_row, column=4).value = data[2]
        
        # Auto-fit the column widths
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
        

# Add pie chart to summary sheet
data = openpyxl.chart.Reference(summary_sheet, min_col=2, min_row=2, max_row=summary_sheet.max_row)
categories = openpyxl.chart.Reference(summary_sheet, min_col=1, min_row=2, max_row=summary_sheet.max_row)
chart = openpyxl.chart.PieChart()
chart.title = "Userwise Data Usage"
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
summary_sheet.add_chart(chart, "F2")
data_labels = DataLabelList()
data_labels.showPercent = True
chart.dataLabels = data_labels
chart.width = 15
chart.height = 10

workbook.save('Usage_information_userwise.xlsx')
