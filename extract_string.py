#Change the workbook name here and run using jupyter notebook 
#This script fetches "rs"/"Rs" entries from all the worksheets of given workbook and writes it into individual txt tiles with sheetnames

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('RS_id_Levels_details.xlsx')

# Iterate over all worksheets in the workbook
for worksheet in workbook.worksheets:
    # Create a new file with the worksheet name
    filename = worksheet.title + '.txt'
    with open(filename, 'w') as f:
        # Iterate over all cells in the current worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                # Check if the cell value starts with "rs" or "Rs"
                if cell.value and cell.value.strip().lower().startswith('rs'):
                    # Write the cell value to the output file
                    print(cell.value, file=f)
                    
